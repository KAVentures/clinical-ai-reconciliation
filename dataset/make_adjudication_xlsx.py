"""Render the blinded clinician packet CSVs into one formatted multi-sheet .xlsx."""
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

HERE = os.path.dirname(os.path.abspath(__file__))
ADJ = os.path.join(HERE, 'adjudication')
OUT_XLSX = os.path.join(ADJ, 'ADJUDICATION_PACKET.xlsx')

ARIAL = 'Arial'
HEADER_FILL = PatternFill('solid', fgColor='1F4E78')
FILL_YELLOW = PatternFill('solid', fgColor='FFF2CC')   # cells the rater edits
EXAMPLE_FILL = PatternFill('solid', fgColor='E2E2E2')
THIN = Side(style='thin', color='BFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# fill-in column -> dropdown options (None = free numeric/text)
VALIDATIONS = {}
for ax in ['accuracy', 'clinical_utility', 'source_quality', 'completeness', 'verifiability']:
    VALIDATIONS[f'rubric_A_{ax}'] = ['1', '2', '3', '4']
    VALIDATIONS[f'rubric_B_{ax}'] = ['1', '2', '3', '4']
    VALIDATIONS[f'prefer_{ax}'] = ['A', 'B', 'tie']
for s in ['A', 'B']:
    VALIDATIONS[f'adj_{s}_factual_correctness'] = ['correct', 'minor_error', 'major_error']
    VALIDATIONS[f'adj_{s}_material_omission'] = ['none', 'minor', 'major']
    VALIDATIONS[f'adj_{s}_potential_harm'] = ['none', 'low', 'moderate', 'high']
    VALIDATIONS[f'adj_{s}_appropriately_qualified'] = ['yes', 'partially', 'no']
    VALIDATIONS[f'adj_{s}_citation_support'] = ['supported', 'partial', 'unsupported', 'no_citations']
VALIDATIONS['overall_preferred_answer'] = ['A', 'B', 'tie']
VALIDATIONS['rater_notes'] = None

EXAMPLE = {  # realistic example values (illustration only)
    'item_id': 'EXAMPLE', 'specialty': 'Cardiology',
    'clinical_question': '(example) In a 68-year-old with new AF and CrCl 40, which DOAC and dose?',
    'answer_A': '(example answer A text — a full clinical answer would appear here)',
    'answer_B': '(example answer B text — a full clinical answer would appear here)',
    'rubric_A_accuracy': 4, 'rubric_A_clinical_utility': 4, 'rubric_A_source_quality': 3,
    'rubric_A_completeness': 4, 'rubric_A_verifiability': 3,
    'rubric_B_accuracy': 3, 'rubric_B_clinical_utility': 3, 'rubric_B_source_quality': 2,
    'rubric_B_completeness': 3, 'rubric_B_verifiability': 2,
    'prefer_accuracy': 'A', 'prefer_clinical_utility': 'A', 'prefer_source_quality': 'A',
    'prefer_completeness': 'A', 'prefer_verifiability': 'A',
    'adj_A_factual_correctness': 'correct', 'adj_A_material_omission': 'none',
    'adj_A_potential_harm': 'none', 'adj_A_appropriately_qualified': 'yes',
    'adj_A_citation_support': 'supported',
    'adj_B_factual_correctness': 'minor_error', 'adj_B_material_omission': 'minor',
    'adj_B_potential_harm': 'low', 'adj_B_appropriately_qualified': 'partially',
    'adj_B_citation_support': 'partial',
    'overall_preferred_answer': 'A', 'rater_notes': 'B omits dose reduction for renal function.',
}


def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name=ARIAL, size=10, bold=True, color='FFFFFF')
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        cell.border = BORDER


def main():
    ratings = pd.read_csv(os.path.join(ADJ, 'ratings_template.csv'), dtype=str).fillna('')
    scales = pd.read_csv(os.path.join(ADJ, 'scales.csv'))
    instr = pd.read_csv(os.path.join(ADJ, 'instructions.csv'))
    cols = list(ratings.columns)
    fixed = ['item_id', 'specialty', 'clinical_question', 'answer_A', 'answer_B']
    fill_cols = [c for c in cols if c not in fixed]

    wb = Workbook()

    # ---- Sheet 1: Start here ----
    s1 = wb.active
    s1.title = 'Start here'
    s1['A1'] = 'Clinician rating & adjudication — instructions'
    s1['A1'].font = Font(name=ARIAL, size=14, bold=True)

    # rater identification (so multiple returned files can be told apart) — yellow input cells
    s1['A3'] = 'Your initials:'; s1['A4'] = 'Date started:'; s1['A5'] = 'Specialty:'
    for rr in (3, 4, 5):
        s1.cell(row=rr, column=1).font = Font(name=ARIAL, size=10, bold=True)
        inp = s1.cell(row=rr, column=2)
        inp.fill = FILL_YELLOW
        inp.border = BORDER
        inp.font = Font(name=ARIAL, size=10)
    s1.cell(row=3, column=3, value='← please fill these three before you start').font = \
        Font(name=ARIAL, size=9, italic=True, color='808080')

    s1['A7'] = ('This is the only file you need. Please complete the YELLOW columns on the "Ratings" sheet. '
                'Allowed values for each column are on the "Scales" sheet and appear as drop-downs. Do not edit '
                'the question or answer text. You are blinded to which AI system wrote each answer. Save and '
                'return this single file when done (you may do it over several sittings).')
    s1['A7'].alignment = Alignment(wrap_text=True, vertical='top')
    s1.merge_cells('A7:H10')
    r = 12
    for _, row in instr.iterrows():
        c = s1.cell(row=r, column=1, value=f"{row['step']}. {row['instruction']}"
                    if str(row['step']).isdigit() else row['instruction'])
        c.font = Font(name=ARIAL, size=10)
        c.alignment = Alignment(wrap_text=True, vertical='top')
        s1.merge_cells(start_row=r, start_column=1, end_row=r + 1, end_column=8)
        r += 3
    s1.column_dimensions['A'].width = 22
    for col in 'BCDEFGH':
        s1.column_dimensions[col].width = 14

    # ---- Sheet 2: Scales ----
    s2 = wb.create_sheet('Scales')
    s2['A1'] = 'Fill-in column'; s2['B1'] = 'Allowed values'
    style_header(s2, 2)
    for i, (_, row) in enumerate(scales.iterrows(), start=2):
        s2.cell(row=i, column=1, value=row['fill_in_column']).font = Font(name=ARIAL, size=10, bold=True)
        s2.cell(row=i, column=2, value=row['allowed_values']).font = Font(name=ARIAL, size=10)
        for c in (1, 2):
            s2.cell(row=i, column=c).border = BORDER
            s2.cell(row=i, column=c).alignment = Alignment(wrap_text=True, vertical='top')
    s2.column_dimensions['A'].width = 34
    s2.column_dimensions['B'].width = 70
    s2.freeze_panes = 'A2'

    # ---- Sheet 3: Ratings ----
    s3 = wb.create_sheet('Ratings')
    for j, col in enumerate(cols, start=1):
        s3.cell(row=1, column=j, value=col)
    style_header(s3, len(cols))

    def write_row(xlrow, data, example=False):
        for j, col in enumerate(cols, start=1):
            v = data.get(col, '')
            if col.startswith('rubric_') and str(v).strip().isdigit():
                v = int(v)
            cell = s3.cell(row=xlrow, column=j, value=v)
            cell.font = Font(name=ARIAL, size=10, italic=example)
            cell.border = BORDER
            cell.alignment = Alignment(wrap_text=col in ('clinical_question', 'answer_A', 'answer_B', 'rater_notes'),
                                       vertical='top', horizontal='center' if col in VALIDATIONS and col != 'rater_notes' else 'left')
            if example:
                cell.fill = EXAMPLE_FILL
            elif col in fill_cols:
                cell.fill = FILL_YELLOW

    write_row(2, EXAMPLE, example=True)
    s3.cell(row=2, column=1, value='EXAMPLE — delete this row').font = Font(name=ARIAL, size=10, bold=True, italic=True)
    for i, (_, row) in enumerate(ratings.iterrows(), start=3):
        write_row(i, row.to_dict())

    # widths
    widths = {'item_id': 12, 'specialty': 18, 'clinical_question': 55, 'answer_A': 65, 'answer_B': 65}
    for j, col in enumerate(cols, start=1):
        s3.column_dimensions[get_column_letter(j)].width = widths.get(col, 15)
    s3.freeze_panes = 'B2'   # keep header row + item_id column visible

    # data validation dropdowns over the data rows (example + 79 items -> rows 2..N+2)
    last = len(ratings) + 2
    for j, col in enumerate(cols, start=1):
        opts = VALIDATIONS.get(col)
        if not opts:
            continue
        dv = DataValidation(type='list', formula1='"' + ','.join(opts) + '"', allow_blank=True, showDropDown=False)
        s3.add_data_validation(dv)
        L = get_column_letter(j)
        dv.add(f'{L}3:{L}{last}')

    wb.save(OUT_XLSX)
    print(f"wrote {OUT_XLSX}  ({len(ratings)} items, {len(fill_cols)} fill-in columns, 3 sheets)")


if __name__ == '__main__':
    main()
