"""Build the physician 2x2-completion study: two genuinely separate rating instruments + reviewer packets.

Completes the local factorial:
                       Pairwise        Absolute rubric
    Physicians         new A'          new D
    LLM judges         existing B      existing C

ABSOLUTE-RUBRIC arm (cell D): Nature FORMAT - ONE blinded answer per row, absolute 1-4, the competing
answer is NEVER shown. Real-POCQi CONTENT + the SAME five axes as the LLM cells. The ~80 (question,opponent)
pairs each create TWO response-items (OE answer + opponent answer), i.e. ~160 response-items before
replication (answers are NOT deduplicated - each pair contributes its own two items).

PAIRWISE arm (cell A'): both blinded answers, A/B randomised independently per item AND per reviewer, five
A/B/tie axis preferences + overall preference.

Outputs (dataset/physician/):
  PHYSICIAN_ABSOLUTE_RUBRIC.xlsx   master template, 158 response-items (Instructions/Rubric anchors/Worked
                                   example/Data dictionary/Ratings)
  PHYSICIAN_PAIRWISE.xlsx          master template, 79 items (Instructions/Worked example/Data dictionary/Ratings)
  packets/PHYSICIAN_*__REV-NN.xlsx reviewer-specific packets (15-25 items; >=2 ratings/item; a reviewer never
                                   sees the same question in both arms or two answers to the same question)
  author_only/  physician_sample.csv, rubric_response_manifest.csv, pairwise_item_manifest.csv,
                pairwise_packet_blinding.csv, assignment_manifest.csv, data_dictionary.csv, AUTHOR_README.md

Blinding: provider names removed and OpenEvidence brand/domain scrubbed from answer text (citations kept).
Deterministic (seed 62). No API calls. Reads dataset/human_study_sample.csv + data/{questions,answers}.parquet.
"""
import os
import re
import sys
import shutil
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.join(HERE, '..', '..')
DATA = os.path.join(ROOT, 'data')
OUTJ = os.path.join(ROOT, 'judge', 'out')
AUTHOR = os.path.join(HERE, 'author_only')
PKT = os.path.join(HERE, 'packets')
sys.path.insert(0, os.path.join(ROOT, 'judge'))
from rubric_anchors import AXES, AXIS_DEF, ANCHORS   # single canonical rubric (also used by grade_expanded.py)
from blinding import render_blinded_answer           # single canonical answer rendering (also used by grade_expanded.py)
OE = 'openevidence'
FRONTIER = ['gpt-5.5', 'claude-opus-4-8', 'gemini-3.1-pro']
SEED = 62
TARGET_PAIRS = 80             # question x opponent pairs (prespecified 60-100); one per distinct question

ARIAL = 'Arial'
HEADER_FILL = PatternFill('solid', fgColor='1F4E78')
YELLOW = PatternFill('solid', fgColor='FFF2CC')
LOCKFILL = PatternFill('solid', fgColor='F2F2F2')
THIN = Side(style='thin', color='BFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COMPETENCE, CONFIDENCE, PREF = ['yes', 'partly', 'no'], ['high', 'moderate', 'low'], ['A', 'B', 'tie']
EXTERNAL_NOTE = ("Ratings must represent your own final clinical judgment. External clinical resources may be "
                 "consulted when needed, but do not delegate completion of the rating form to another person or system.")


# ---------------------------------------------------------------- sheet helpers
def hdr(ws, headers, row):
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = Font(name=ARIAL, size=10, bold=True, color='FFFFFF')
        c.fill = HEADER_FILL; c.border = BORDER
        c.alignment = Alignment(vertical='center', wrap_text=True, horizontal='center')


def instructions_sheet(wb, title, task_lines):
    ws = wb.create_sheet('Instructions')
    ws['A1'] = title; ws['A1'].font = Font(name=ARIAL, size=14, bold=True)
    ws['A3'] = 'Reviewer initials:'; ws['A4'] = 'Date started:'; ws['A5'] = 'Your specialty / clinical background:'
    for rr in (3, 4, 5):
        ws.cell(row=rr, column=1).font = Font(name=ARIAL, size=10, bold=True)
        inp = ws.cell(row=rr, column=2); inp.fill = YELLOW; inp.border = BORDER
        inp.font = Font(name=ARIAL, size=10); inp.protection = Protection(locked=False)
    r = 7
    for line in task_lines + ['', EXTERNAL_NOTE,
                              'You are blinded to which AI system produced each answer. Judge only the content.',
                              'Fill only the YELLOW cells on the "Ratings" sheet (drop-downs enforce allowed values). '
                              'Question and answer text are locked - please do not alter them.',
                              'A progress counter (rows scored / remaining) is at the top of the "Ratings" sheet.']:
        c = ws.cell(row=r, column=1, value=line)
        c.font = Font(name=ARIAL, size=10, bold=line.endswith(':'))
        c.alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells(start_row=r, start_column=1, end_row=r + 1, end_column=9)
        r += 3
    ws.column_dimensions['A'].width = 26
    for col in 'BCDEFGHI':
        ws.column_dimensions[col].width = 12
    ws.protection.sheet = True


def anchors_sheet(wb):
    ws = wb.create_sheet('Rubric anchors')
    ws['A1'] = ('Axis definitions are verbatim from the study rubric (judge/grade.py); score 1-4 = '
                'unacceptable / marginal / good / excellent, elaborated per axis below.')
    ws['A1'].font = Font(name=ARIAL, size=10, italic=True)
    ws.merge_cells('A1:F1')
    hdr(ws, ['Axis', 'Definition (grade.py)', 'Score 1 (unacceptable)', 'Score 2 (marginal)', 'Score 3 (good)', 'Score 4 (excellent)'], 2)
    for i, ax in enumerate(AXES, start=3):
        ws.cell(row=i, column=1, value=ax.replace('_', ' ')).font = Font(name=ARIAL, size=10, bold=True)
        ws.cell(row=i, column=2, value=AXIS_DEF[ax]).font = Font(name=ARIAL, size=10)
        for k in (1, 2, 3, 4):
            ws.cell(row=i, column=2 + k, value=ANCHORS[ax][k]).font = Font(name=ARIAL, size=10)
        for j in range(1, 7):
            ws.cell(row=i, column=j).alignment = Alignment(wrap_text=True, vertical='top'); ws.cell(row=i, column=j).border = BORDER
    ws.column_dimensions['A'].width = 16; ws.column_dimensions['B'].width = 30
    for col in 'CDEF':
        ws.column_dimensions[col].width = 40
    ws.freeze_panes = 'A3'; ws.protection.sheet = True


def worked_example_sheet(wb, headers, example, note):
    ws = wb.create_sheet('Worked example')
    ws['A1'] = 'Worked example - for reference only (not part of your ratings)'
    ws['A1'].font = Font(name=ARIAL, size=12, bold=True)
    hdr(ws, headers, 3)
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=j, value=example.get(h, ''))
        c.font = Font(name=ARIAL, size=10); c.border = BORDER
        c.alignment = Alignment(wrap_text=h in ('clinical_question', 'blinded_answer', 'answer_A', 'answer_B', 'optional_comment'), vertical='top')
    ws.cell(row=6, column=1, value=note).font = Font(name=ARIAL, size=10, italic=True)
    ws.merge_cells(start_row=6, start_column=1, end_row=7, end_column=len(headers))
    ws.cell(row=6, column=1).alignment = Alignment(wrap_text=True, vertical='top')
    for j, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(j)].width = 60 if ('answer' in h or h == 'clinical_question') else 16
    ws.protection.sheet = True


def data_dictionary_sheet(wb, entries):
    ws = wb.create_sheet('Data dictionary')
    hdr(ws, ['Field', 'Allowed values', 'Description'], 1)
    for i, (f, a, d) in enumerate(entries, start=2):
        ws.cell(row=i, column=1, value=f).font = Font(name=ARIAL, size=10, bold=True)
        ws.cell(row=i, column=2, value=a).font = Font(name=ARIAL, size=10)
        ws.cell(row=i, column=3, value=d).font = Font(name=ARIAL, size=10)
        for j in (1, 2, 3):
            ws.cell(row=i, column=j).alignment = Alignment(wrap_text=True, vertical='top'); ws.cell(row=i, column=j).border = BORDER
    ws.column_dimensions['A'].width = 28; ws.column_dimensions['B'].width = 34; ws.column_dimensions['C'].width = 60
    ws.freeze_panes = 'A2'; ws.protection.sheet = True


def ratings_sheet(wb, headers, rows, fill_cols, validations, read_cols):
    ws = wb.create_sheet('Ratings')
    n = len(rows); first, last = 3, n + 2
    # a row counts COMPLETE only when EVERY required field (all fill cols except the optional comment) is
    # filled: SUMPRODUCT of per-column non-blank indicators = count of fully-completed rows.
    required = [c for c in fill_cols if c != 'optional_comment']
    reqL = [get_column_letter(headers.index(c) + 1) for c in required]
    completed = '=SUMPRODUCT(' + ','.join(f'--({L}{first}:{L}{last}<>"")' for L in reqL) + ')'
    ws['A1'] = 'Rows complete:'; ws['B1'] = completed
    ws['C1'] = 'Remaining:'; ws['D1'] = f'={n}-B1'
    ws['E1'] = f'Total items: {n}'
    for cell in ('A1', 'B1', 'C1', 'D1', 'E1'):
        ws[cell].font = Font(name=ARIAL, size=10, bold=True, color='1F4E78')
    hdr(ws, headers, 2)
    for i, row in enumerate(rows, start=first):
        for j, h in enumerate(headers, 1):
            c = ws.cell(row=i, column=j, value=row.get(h, ''))
            c.alignment = Alignment(wrap_text=h in ('clinical_question', 'blinded_answer', 'answer_A', 'answer_B', 'optional_comment'),
                                    vertical='top', horizontal='center' if h in validations else 'left')
            c.font = Font(name=ARIAL, size=10); c.border = BORDER
            if h in fill_cols:
                c.fill = YELLOW; c.protection = Protection(locked=False)
            elif h in read_cols:
                c.fill = LOCKFILL
    for j, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(j)].width = 60 if ('answer' in h or h == 'clinical_question') else (12 if h in validations else 16)
    for h, opts in validations.items():
        dv = DataValidation(type='list', formula1='"' + ','.join(opts) + '"', allow_blank=True)
        ws.add_data_validation(dv); L = get_column_letter(headers.index(h) + 1)
        dv.add(f'{L}{first}:{L}{last}')
    ws.freeze_panes = 'B3'; ws.protection.sheet = True


# ---------------------------------------------------------------- workbook builders
RUBRIC_DD = ([('item_id', 'RESP-####', 'blinded response-item id (one AI answer)'),
              ('blinded_question_id', 'QID-###', 'blinded question; SAME id = same clinical question'),
              ('clinical_question', 'text (locked)', 'the point-of-care question'),
              ('blinded_answer', 'text (locked)', 'ONE AI answer; provider hidden; no competing answer shown')] +
             [(f'{ax}_score', '1 / 2 / 3 / 4', f'absolute score on {ax}: {AXIS_DEF[ax]} (anchors on Rubric anchors sheet)') for ax in AXES] +
             [('within_reviewer_competence', 'yes / partly / no', 'is the item within your area of competence'),
              ('reviewer_confidence', 'high / moderate / low', 'your confidence in this rating'),
              ('optional_comment', 'free text', 'optional note / safety concern')])
PAIRWISE_DD = ([('item_id', 'PW-###', 'blinded comparison id'),
                ('blinded_question_id', 'QID-###', 'blinded question; SAME id = same clinical question'),
                ('clinical_question', 'text (locked)', 'the point-of-care question'),
                ('answer_A', 'text (locked)', 'first answer (A/B order randomised per item and reviewer)'),
                ('answer_B', 'text (locked)', 'second answer')] +
               [(f'{ax}_preference', 'A / B / tie', f'which answer is better on {ax}: {AXIS_DEF[ax]}') for ax in AXES] +
               [('overall_preference', 'A / B / tie', 'overall preferred answer'),
                ('within_reviewer_competence', 'yes / partly / no', 'is the item within your area of competence'),
                ('reviewer_confidence', 'high / moderate / low', 'your confidence in this comparison'),
                ('optional_comment', 'free text', 'optional note / safety concern')])


def build_rubric_workbook(rows, path):
    headers = (['item_id', 'blinded_question_id', 'clinical_question', 'blinded_answer'] +
               [f'{ax}_score' for ax in AXES] + ['within_reviewer_competence', 'reviewer_confidence', 'optional_comment'])
    fill_cols = [f'{ax}_score' for ax in AXES] + ['within_reviewer_competence', 'reviewer_confidence', 'optional_comment']
    read_cols = ['item_id', 'blinded_question_id', 'clinical_question', 'blinded_answer']
    validations = {**{f'{ax}_score': ['1', '2', '3', '4'] for ax in AXES},
                   'within_reviewer_competence': COMPETENCE, 'reviewer_confidence': CONFIDENCE}
    wb = Workbook(); wb.remove(wb.active)
    instructions_sheet(wb, 'Physician absolute-rubric evaluation - instructions', [
        'TASK: You will see ONE clinical question and ONE AI-generated answer at a time. Score that answer on '
        'its own merits - there is NO competing answer to compare against.',
        'Give an integer 1-4 on each of the five axes (accuracy, clinical utility, source quality, '
        'completeness, verifiability). 1 = unacceptable, 2 = marginal, 3 = good, 4 = excellent.',
        'Use the axis definitions and per-score anchors on the "Rubric anchors" sheet.',
        'within_reviewer_competence (yes / partly / no); reviewer_confidence (high / moderate / low); optional_comment.'])
    anchors_sheet(wb)
    worked_example_sheet(wb, headers, {
        'item_id': 'RESP-EXAMPLE', 'blinded_question_id': 'QID-EX',
        'clinical_question': '(example) In a 68-year-old with new AF and CrCl 40 mL/min, which DOAC and dose?',
        'blinded_answer': '(example answer text would appear here)',
        'accuracy_score': 3, 'clinical_utility_score': 4, 'source_quality_score': 3, 'completeness_score': 3,
        'verifiability_score': 2, 'within_reviewer_competence': 'yes', 'reviewer_confidence': 'high',
        'optional_comment': 'Correct DOAC choice; dose caveat could be more explicit.'},
        'Score each answer independently on all five axes using the anchors sheet.')
    data_dictionary_sheet(wb, RUBRIC_DD)
    ratings_sheet(wb, headers, rows, fill_cols, validations, read_cols)
    wb.save(path)


def build_pairwise_workbook(items, path, rng):
    """items: list of {item_id, blinded_question_id, clinical_question, oe_answer, opp_answer, opponent, question_id}.
    Randomises A/B per item using rng; returns blinding list [(item_id, slot_A_provider, slot_B_provider)]."""
    headers = (['item_id', 'blinded_question_id', 'clinical_question', 'answer_A', 'answer_B'] +
               [f'{ax}_preference' for ax in AXES] + ['overall_preference', 'within_reviewer_competence', 'reviewer_confidence', 'optional_comment'])
    fill_cols = [f'{ax}_preference' for ax in AXES] + ['overall_preference', 'within_reviewer_competence', 'reviewer_confidence', 'optional_comment']
    read_cols = ['item_id', 'blinded_question_id', 'clinical_question', 'answer_A', 'answer_B']
    validations = {**{f'{ax}_preference': PREF for ax in AXES}, 'overall_preference': PREF,
                   'within_reviewer_competence': COMPETENCE, 'reviewer_confidence': CONFIDENCE}
    rows, blinding = [], []
    for it in items:
        oe_is_A = bool(rng.integers(0, 2))
        aA, aB = (it['oe_answer'], it['opp_answer']) if oe_is_A else (it['opp_answer'], it['oe_answer'])
        pA, pB = (OE, it['opponent']) if oe_is_A else (it['opponent'], OE)
        rows.append({'item_id': it['item_id'], 'blinded_question_id': it['blinded_question_id'],
                     'clinical_question': it['clinical_question'], 'answer_A': aA, 'answer_B': aB})
        blinding.append({'item_id': it['item_id'], 'question_id': it['question_id'], 'opponent': it['opponent'],
                         'slot_A_provider': pA, 'slot_B_provider': pB})
    wb = Workbook(); wb.remove(wb.active)
    instructions_sheet(wb, 'Physician pairwise-preference evaluation - instructions', [
        'TASK: You will see ONE clinical question and TWO AI-generated answers, Answer A and Answer B '
        '(order is randomised). For each of the five axes say which is better: A, B, or tie.',
        'Axis definitions (apply the same meaning as the rubric arm):',
        *[f'   - {ax.replace("_", " ")}: {AXIS_DEF[ax]}' for ax in AXES],
        'overall_preference: all things considered, which answer would you rather give a colleague - A / B / tie.',
        'within_reviewer_competence (yes / partly / no); reviewer_confidence (high / moderate / low); optional_comment.'])
    worked_example_sheet(wb, headers, {
        'item_id': 'PW-EXAMPLE', 'blinded_question_id': 'QID-EX',
        'clinical_question': '(example) In a 68-year-old with new AF and CrCl 40 mL/min, which DOAC and dose?',
        'answer_A': '(example Answer A text)', 'answer_B': '(example Answer B text)',
        'accuracy_preference': 'A', 'clinical_utility_preference': 'A', 'source_quality_preference': 'tie',
        'completeness_preference': 'A', 'verifiability_preference': 'B', 'overall_preference': 'A',
        'within_reviewer_competence': 'yes', 'reviewer_confidence': 'moderate',
        'optional_comment': 'A gives the renal dose adjustment; B is better referenced.'},
        'Choose A, B, or tie on each axis, then an overall preference.')
    data_dictionary_sheet(wb, PAIRWISE_DD)
    ratings_sheet(wb, headers, rows, fill_cols, validations, read_cols)
    wb.save(path)
    return blinding


# ---------------------------------------------------------------- assignment (balanced incomplete block)
def assign_bib(items, n_doctors, max_per_doctor, target_ratings, seed, prefix='DR'):
    """ONE scalable balanced-incomplete-block assignment; the doctor count only sets replication.

    items: list of dicts, each {'id', 'question_id', 'balance': (category values to balance across doctors)}.
    Rules: every item rated >= 1 (coverage first), then 2nd/3rd ratings up to target_ratings while capacity
    remains; a doctor never sees the same clinical question twice; loads capped at max_per_doctor; each doctor
    balanced on the 'balance' categories (e.g. provider OE/competitor, opponent model, specialty).
    Returns (rows, doctors, load, raters) where raters[id] = list of doctors who rated it."""
    from collections import Counter
    docs = [f'{prefix}-{k + 1:02d}' for k in range(n_doctors)]
    seen = {d: set() for d in docs}; load = {d: 0 for d in docs}
    cat = {d: Counter() for d in docs}                 # how many of each balance-value doctor d already has
    raters = {it['id']: [] for it in items}
    rng = np.random.default_rng(seed); rows = []
    for rnd in range(target_ratings):                  # pass 0 = coverage; passes 1.. = replication
        for it in [items[i] for i in rng.permutation(len(items))]:
            if len(raters[it['id']]) > rnd or len(raters[it['id']]) >= target_ratings:
                continue
            elig = [d for d in docs if it['question_id'] not in seen[d]
                    and load[d] < max_per_doctor and d not in raters[it['id']]]
            if not elig:
                continue                               # no doctor can take it under the constraints/capacity
            d = min(elig, key=lambda d: (load[d], sum(cat[d][b] for b in it['balance']), rng.random()))
            seen[d].add(it['question_id']); load[d] += 1
            for b in it['balance']:
                cat[d][b] += 1
            raters[it['id']].append(d)
            rows.append({'doctor_id': d, 'item_id': it['id'], 'question_id': it['question_id'],
                         'rating_round': rnd + 1, 'presentation_order': load[d]})
    return rows, docs, load, raters


# ---------------------------------------------------------------- sample (direct question x opponent)
def select_pairs():
    """Stratified sample built DIRECTLY at the (question, opponent) level (not axis-level rows deduplicated).
    Strata = opponent x outcome-class, where outcome-class in {flip, tie_near, agree}, so tie/near-tie pairs
    are represented (the study is disagreement-ENRICHED but not disagreement-only)."""
    df = pd.read_csv(os.path.join(OUTJ, 'instrument_disagreement.csv'))
    spec = pd.read_parquet(os.path.join(DATA, 'questions.parquet')).set_index('question_id')['specialty']
    df['abs_margin'] = pd.to_numeric(df.rubric_margin, errors='coerce').abs()   # ABSOLUTE margin
    df['flip'] = df.instrument_flip_llm.astype(str).str.lower().eq('true')
    df['is_tie'] = df.rubric_winner.astype(str).eq('tie')
    pairs = (df.groupby(['question_id', 'opponent'])
             .agg(n_axes=('axis', 'count'), n_flip=('flip', 'sum'), n_tie=('is_tie', 'sum'),
                  mean_abs_margin=('abs_margin', 'mean')).reset_index())
    dfm = df.dropna(subset=['abs_margin'])
    sig = dfm.loc[dfm.groupby(['question_id', 'opponent']).abs_margin.idxmax(), ['question_id', 'opponent', 'axis']]
    pairs = pairs.merge(sig, on=['question_id', 'opponent'], how='left').rename(columns={'axis': 'signal_axis'})
    pairs['specialty'] = pairs.question_id.map(spec)
    # require a defined margin (i.e. LLM rubric data exists for the pair -> cell C is estimable)
    pairs = pairs.dropna(subset=['mean_abs_margin']).reset_index(drop=True)

    def outcome(r):
        if r.n_flip >= 1:
            return 'flip'
        if r.n_tie >= 2 or (pd.notna(r.mean_abs_margin) and r.mean_abs_margin < 0.25):
            return 'tie_near'
        return 'agree'
    pairs['outcome_class'] = pairs.apply(outcome, axis=1)
    # ONE (question, opponent) per clinical question, so the balanced-incomplete-block stays clean: each
    # question contributes exactly two response-items, coverable even by two doctors without repeating a question.
    pairs = pairs.sample(frac=1.0, random_state=SEED).drop_duplicates(subset='question_id', keep='first')

    rng = np.random.default_rng(SEED)
    strata = [(o, c) for o in FRONTIER for c in ['flip', 'tie_near', 'agree']]
    per = max(1, TARGET_PAIRS // len(strata) + 1)
    picks = []
    for (o, c) in strata:
        sub = pairs[(pairs.opponent == o) & (pairs.outcome_class == c)]
        if len(sub):
            picks.append(sub.iloc[rng.choice(len(sub), min(per, len(sub)), replace=False)])
    sel = pd.concat(picks).drop_duplicates(subset=['question_id', 'opponent'])
    if len(sel) > TARGET_PAIRS:
        sel = sel.iloc[rng.choice(len(sel), TARGET_PAIRS, replace=False)]
    return sel.sort_values(['question_id', 'opponent'], kind='stable').reset_index(drop=True)


# ---------------------------------------------------------------- coverage report
def coverage(raters, n_items):
    from collections import Counter
    cov = Counter(len(v) for v in raters.values())
    n_cov = sum(1 for v in raters.values() if v)
    return {'items': n_items, 'covered_>=1': n_cov, 'uncovered_0': n_items - n_cov,
            'ratings_per_item_dist': dict(sorted(cov.items())),
            'min_ratings': min(len(v) for v in raters.values()), 'max_ratings': max(len(v) for v in raters.values())}


# ---------------------------------------------------------------- main
def main():
    import argparse
    ap = argparse.ArgumentParser(description="Scalable physician study builder (rubric arm D is essential; "
                                             "pairwise arm A' is optional). Doctor count only sets replication.")
    ap.add_argument('--doctors', type=int, default=3, help='number of recruited doctors (rubric arm)')
    ap.add_argument('--max-per-doctor', type=int, default=60, help='max rating tasks per doctor')
    ap.add_argument('--target-ratings', type=int, default=2, help='desired ratings per response (>=1 guaranteed if capacity)')
    ap.add_argument('--pairwise', action='store_true', help="ALSO build the optional physician pairwise arm (A')")
    ap.add_argument('--pairwise-doctors', type=int, default=2)
    args = ap.parse_args()

    for d in (AUTHOR, PKT):
        if os.path.isdir(d):
            shutil.rmtree(d)
        os.makedirs(d)
    q = pd.read_parquet(os.path.join(DATA, 'questions.parquet')).set_index('question_id')
    a = pd.read_parquet(os.path.join(DATA, 'answers.parquet'))
    ANS = {(r.question_id, r.provider_key): render_blinded_answer(r.answer_markdown) for _, r in a.iterrows()}

    pairs = select_pairs()
    pairs.insert(0, 'pair_id', [f'PAIR-{i + 1:03d}' for i in range(len(pairs))])
    pairs.to_csv(os.path.join(AUTHOR, 'physician_sample.csv'), index=False)
    qid_map = {qid: f'QID-{i + 1:03d}' for i, qid in enumerate(sorted(set(pairs.question_id)))}

    # ---- ESSENTIAL arm D: 2 single-answer response-items per pair (NOT deduplicated) ----
    rub_rows, rub_manifest, rub_bib = [], [], []
    k = 0
    for p in pairs.itertuples():
        for prov in (OE, p.opponent):
            k += 1; rid = f'RESP-{k:04d}'
            rub_rows.append({'item_id': rid, 'blinded_question_id': qid_map[p.question_id],
                             'clinical_question': q.loc[p.question_id, 'question_text'], 'blinded_answer': ANS[(p.question_id, prov)]})
            rub_manifest.append({'response_id': rid, 'pair_id': p.pair_id, 'blinded_question_id': qid_map[p.question_id],
                                 'question_id': p.question_id, 'true_provider': prov, 'opponent': p.opponent, 'specialty': p.specialty})
            # balance each doctor on provider-type, opponent model, and specialty
            rub_bib.append({'id': rid, 'question_id': p.question_id,
                            'balance': ('OE' if prov == OE else 'competitor', p.opponent, str(p.specialty))})
    build_rubric_workbook(rub_rows, os.path.join(HERE, 'PHYSICIAN_ABSOLUTE_RUBRIC.xlsx'))
    pd.DataFrame(rub_manifest).to_csv(os.path.join(AUTHOR, 'rubric_response_manifest.csv'), index=False)

    # ---- rubric assignment (balanced incomplete block; doctor count sets replication) ----
    asg, docs, load, raters = assign_bib(rub_bib, args.doctors, args.max_per_doctor, args.target_ratings, SEED, 'DR')
    pd.DataFrame(asg).to_csv(os.path.join(AUTHOR, 'assignment_manifest.csv'), index=False)
    cov = coverage(raters, len(rub_rows))

    # ---- rubric reviewer packets (one per doctor; randomized item order) ----
    rub_by_id = {r['item_id']: r for r in rub_rows}
    asg_df = pd.DataFrame(asg); n_files = 0
    for i, d in enumerate(docs):
        ids = asg_df[asg_df.doctor_id == d].item_id.tolist()
        if not ids:
            continue
        prng = np.random.default_rng(SEED + 1000 + i)
        rows = [rub_by_id[x] for x in ids]
        rows = [rows[j] for j in prng.permutation(len(rows))]
        build_rubric_workbook(rows, os.path.join(PKT, f'PHYSICIAN_ABSOLUTE_RUBRIC__{d}.xlsx'))
        n_files += 1

    # ---- OPTIONAL arm A': physician pairwise (only with --pairwise) ----
    pw_items = [{'item_id': f'PW-{i + 1:03d}', 'pair_id': p.pair_id, 'blinded_question_id': qid_map[p.question_id],
                 'clinical_question': q.loc[p.question_id, 'question_text'], 'question_id': p.question_id,
                 'opponent': p.opponent, 'specialty': p.specialty,
                 'oe_answer': ANS[(p.question_id, OE)], 'opp_answer': ANS[(p.question_id, p.opponent)]}
                for i, p in enumerate(pairs.itertuples())]
    master_blinding = build_pairwise_workbook(pw_items, os.path.join(HERE, 'PHYSICIAN_PAIRWISE.xlsx'), np.random.default_rng(SEED))
    pd.DataFrame(master_blinding).to_csv(os.path.join(AUTHOR, 'pairwise_item_manifest.csv'), index=False)
    pw_cov = None
    if args.pairwise:
        pw_bib = [{'id': it['item_id'], 'question_id': it['question_id'], 'balance': (it['opponent'], str(it['specialty']))} for it in pw_items]
        pasg, pdocs, pload, praters = assign_bib(pw_bib, args.pairwise_doctors, args.max_per_doctor, args.target_ratings, SEED + 7, 'DP')
        pd.DataFrame(pasg).to_csv(os.path.join(AUTHOR, 'pairwise_assignment_manifest.csv'), index=False)
        pw_by_id = {it['item_id']: it for it in pw_items}; pasg_df = pd.DataFrame(pasg); pblind = []
        for i, d in enumerate(pdocs):
            ids = pasg_df[pasg_df.doctor_id == d].item_id.tolist()
            if not ids:
                continue
            prng = np.random.default_rng(SEED + 2000 + i)
            items = [pw_by_id[x] for x in ids]; items = [items[j] for j in prng.permutation(len(items))]
            bl = build_pairwise_workbook(items, os.path.join(PKT, f'PHYSICIAN_PAIRWISE__{d}.xlsx'), prng)
            for b in bl:
                b['doctor_id'] = d; pblind.append(b)
            n_files += 1
        pd.DataFrame(pblind).to_csv(os.path.join(AUTHOR, 'pairwise_packet_blinding.csv'), index=False)
        pw_cov = coverage(praters, len(pw_items))

    write_data_dictionary(os.path.join(AUTHOR, 'data_dictionary.csv'))
    write_author_readme(os.path.join(AUTHOR, 'AUTHOR_README.md'), len(pairs), len(rub_rows), args, docs, load, cov, n_files, pw_cov)

    print(f"pairs (distinct questions): {len(pairs)}  (opponents {pairs.opponent.value_counts().to_dict()})")
    print(f"ESSENTIAL rubric arm D    : {len(rub_rows)} response-items (= {len(pairs)} x 2) -> PHYSICIAN_ABSOLUTE_RUBRIC.xlsx")
    print(f"  {args.doctors} doctors, max {args.max_per_doctor}/doctor, target {args.target_ratings} ratings/response")
    print(f"  coverage: {cov['covered_>=1']}/{cov['items']} covered; ratings/response dist {cov['ratings_per_item_dist']}; "
          f"doctor loads {sorted(load.values())}")
    if cov['uncovered_0']:
        print(f"  ** {cov['uncovered_0']} responses UNCOVERED - add doctors or raise --max-per-doctor **")
    print(f"pairwise arm A' (optional): {'BUILT ('+str(pw_cov['covered_>=1'])+' covered)' if pw_cov else 'master only; pass --pairwise to build packets'}")
    print(f"reviewer packets          : {n_files} in {PKT}")


def write_data_dictionary(path):
    rows = [{'workbook': 'rubric', 'field': f, 'allowed_values': a, 'description': d} for f, a, d in RUBRIC_DD]
    rows += [{'workbook': 'pairwise', 'field': f, 'allowed_values': a, 'description': d} for f, a, d in PAIRWISE_DD]
    rows += [{'workbook': 'author_only', 'field': f, 'allowed_values': a, 'description': d} for f, a, d in [
        ('true_provider / slot_*_provider', 'openevidence|gpt-5.5|claude-opus-4-8|gemini-3.1-pro', 'un-blinding map (never sent to reviewers)'),
        ('pair_id', 'PAIR-###', 'the (question, opponent) pair; two RESP items (and one optional PW item) share a pair_id'),
        ('doctor_id', 'DR-## (rubric) / DP-## (optional pairwise)', 'assigned physician (placeholder; map via reviewer_allocation.csv)'),
        ('rating_round', 'int', '1 = coverage rating; 2,3 = replication ratings'),
        ('presentation_order', 'int', 'order the item is shown to that doctor')]]
    pd.DataFrame(rows).to_csv(path, index=False)


def write_author_readme(path, n_pairs, n_rub, args, docs, load, cov, n_files, pw_cov):
    open(path, 'w').write(f"""# Author-only files — DO NOT SEND TO REVIEWERS

Un-blind the study and define the assignment. Keep private (this folder is git-ignored).

- `physician_sample.csv` — the {n_pairs} (question, opponent) pairs (one per distinct clinical question) + strata.
- `rubric_response_manifest.csv` — RESP-#### -> pair_id, question_id, true_provider, opponent, specialty
  ({n_rub} response-items = {n_pairs} pairs x 2, NOT deduplicated). This un-blinds the ESSENTIAL rubric arm D.
- `assignment_manifest.csv` — balanced-incomplete-block: doctor_id x response x rating_round.
  {args.doctors} doctors, max {args.max_per_doctor}/doctor, target {args.target_ratings} ratings/response.
  Coverage {cov['covered_>=1']}/{cov['items']}; ratings/response {cov['ratings_per_item_dist']}; loads {sorted(load.values())}.
  Every response rated >=1; a doctor never sees a clinical question twice; provider/opponent/specialty balanced.
- `pairwise_item_manifest.csv` — optional arm A' master A/B mapping. `pairwise_assignment_manifest.csv` /
  `pairwise_packet_blinding.csv` exist only if built with --pairwise (A/B re-randomised per doctor -> un-blind
  pairwise with the PACKET blinding file, not the master).
- `data_dictionary.csv` — all fields + allowed values (also a sheet inside each workbook).

Seed = {SEED}. One scalable design: doctor count only sets replication (`--doctors`, `--max-per-doctor`,
`--target-ratings`); rebuild to rescale, no redesign.

## To run
1. `python3 allocate_reviewers.py --roster <roster.csv>` maps real physicians -> one arm + one doctor_id each.
2. `../packets/` holds {n_files} per-doctor files (PHYSICIAN_ABSOLUTE_RUBRIC__DR-NN.xlsx). Send each doctor ONLY theirs.
   Top-level PHYSICIAN_*.xlsx are MASTER templates (all items) - reference only.
3. On return, un-blind rubric via rubric_response_manifest.csv, then run `analyze_returns.py`:
   the ESSENTIAL analysis is human cell D (individual 1-4 scores; provider = main predictor; DOCTOR as a
   FIXED effect since 2-3 doctors are not a random sample; question clustering; axis-specific;
   competence/confidence sensitivity), and D vs matched expanded-anchor C on the OE-superiority scale.

## Caveats to record
- Run cell C as the EXPANDED-anchor regrade (grade_expanded.py) so C and D share rubric wording AND rendering
  (both use judge/blinding.render_blinded_answer). The original B/C used raw text - disclose or regrade.
- Specialty matching unavailable: retain all ratings; sensitivity analysis excluding out-of-competence /
  low-confidence rows.
- With one rating per response (few doctors) you cannot estimate physician inter-rater reliability; the study
  still fills cell D. Add doctors -> more replication (2nd/3rd ratings), same instrument and analysis.
""")


if __name__ == '__main__':
    main()
