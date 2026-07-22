"""Automated assertions for the physician study design (coverage, reviewer loads, no repeated questions).

Deterministic (seed 62); operates on the in-memory sampler/assigner plus the generated workbooks.
Run after build_physician_study.py:  python3 test_physician_study.py
"""
import glob
from collections import Counter
import build_physician_study as B


def test_sample_pair_level_and_enriched():
    pairs = B.select_pairs()
    assert 60 <= len(pairs) <= 100, f"pairs {len(pairs)} outside prespecified 60-100"
    assert set(pairs.opponent) == set(B.FRONTIER), "not all three opponents represented"
    classes = set(pairs.outcome_class)
    assert {'flip', 'tie_near', 'agree'} <= classes, f"missing outcome classes: {classes}"
    assert (pairs.outcome_class == 'tie_near').sum() >= 5, "tie/near-tie cases not included"
    # every pair is a distinct (question, opponent)
    assert not pairs.duplicated(subset=['question_id', 'opponent']).any()
    # ABS-MARGIN guard: a tie/near-tie must have a genuinely small |margin| OR >=2 tie axes - a strong
    # (possibly negative) margin must NOT be mislabelled near-tie. Catches the missing .abs() bug.
    tn = pairs[pairs.outcome_class == 'tie_near']
    assert ((tn.mean_abs_margin < 0.25) | (tn.n_tie >= 2)).all(), "tie_near mislabelled (abs-margin bug?)"
    assert (pairs.mean_abs_margin >= 0).all(), "mean_abs_margin has negatives -> .abs() missing"
    return pairs


def _item_lists(pairs):
    rub, pw, k = [], [], 0
    for r in pairs.itertuples():
        for _ in range(2):                      # two response-items per pair, NOT deduplicated
            k += 1; rub.append((f'RESP-{k:04d}', r.question_id))
    for i, r in enumerate(pairs.itertuples()):
        pw.append((f'PW-{i + 1:03d}', r.question_id))
    return rub, pw


def test_rubric_is_two_per_pair():
    pairs = B.select_pairs()
    rub, pw = _item_lists(pairs)
    assert len(rub) == 2 * len(pairs), "rubric response-items must be 2 x pairs (no dedup)"
    assert len(pw) == len(pairs)


def _rubric_bib_items(pairs):
    return [{'id': f'{r.question_id}|{p}', 'question_id': r.question_id,
             'balance': (p, r.opponent, str(r.specialty))}
            for r in pairs.itertuples() for p in ('OE', 'competitor')]


def test_bib_scales_with_doctor_count():
    """One scalable balanced-incomplete-block design: coverage, no repeated question, provider balance, and
    loads within cap - across different doctor counts. Doctor count only changes replication."""
    pairs = B.select_pairs()
    items = _rubric_bib_items(pairs)
    assert len(items) == 2 * len(pairs)
    idprov = {it['id']: it['balance'][0] for it in items}
    for n_doc, maxpd, target in [(2, 80, 1), (3, 60, 2), (4, 60, 3)]:
        rows, docs, load, raters = B.assign_bib(items, n_doc, maxpd, target, B.SEED)
        # capacity covers every response at least once
        assert min(len(v) for v in raters.values()) >= 1, f"{n_doc} doctors: an item is uncovered"
        assert max(len(v) for v in raters.values()) <= target, f"{n_doc} doctors: over target ratings"
        # a doctor never sees the same clinical question twice
        assert max(Counter((r['doctor_id'], r['question_id']) for r in rows).values()) == 1
        # loads within the per-doctor cap
        assert max(load.values()) <= maxpd
        # each doctor sees roughly equal OE and competitor answers
        pc = {d: {'OE': 0, 'competitor': 0} for d in docs}
        for r in rows:
            pc[r['doctor_id']][idprov[r['item_id']]] += 1
        for d in docs:
            assert abs(pc[d]['OE'] - pc[d]['competitor']) <= 6, f"{n_doc} doctors: provider imbalance for {d}"


def test_workbooks_present_scrubbed_no_example_row():
    import re
    from openpyxl import load_workbook
    files = ['PHYSICIAN_ABSOLUTE_RUBRIC.xlsx', 'PHYSICIAN_PAIRWISE.xlsx'] + glob.glob('packets/*.xlsx')
    for f in files:
        wb = load_workbook(f)
        assert 'Ratings' in wb.sheetnames and 'Data dictionary' in wb.sheetnames, f"{f}: missing required sheet"
        rs = wb['Ratings']
        assert str(rs['B1'].value).startswith('=SUMPRODUCT'), f"{f}: completion counter not all-fields"
        # no EXAMPLE row in the live data
        assert not any(str(rs.cell(r, 1).value).startswith('EXAMPLE') for r in range(3, rs.max_row + 1)), f"{f}: example row in live data"
        # provider brand scrubbed
        for sh in wb.sheetnames:
            for row in wb[sh].iter_rows(values_only=True):
                for v in row:
                    assert not (isinstance(v, str) and re.search('openevidence', v, re.I)), f"{f}: brand leak"
    rub = load_workbook('PHYSICIAN_ABSOLUTE_RUBRIC.xlsx')['Ratings']
    hh = [rub.cell(2, c).value for c in range(1, rub.max_column + 1)]
    assert 'blinded_answer' in hh and 'answer_B' not in hh, "rubric must show ONE answer, no competing answer"


def test_common_estimands():
    import numpy as np
    import analyze_returns as A
    # pairwise OE win-score
    assert A.win_score(['oe', 'oe', 'opp', 'tie']) == 0.625
    # rubric probability-of-superiority from INDEPENDENT score sets
    assert A.prob_superiority([4, 4, 3], [2, 2, 2]) == 1.0     # OE always higher
    assert A.prob_superiority([2, 2], [2, 2]) == 0.5           # all ties -> 0.5
    assert A.prob_superiority([3, 4], [2, 4]) == 0.625         # (2 gt + 0.5*1 eq)/4
    assert np.isnan(A.prob_superiority([], [1, 2]))            # empty side -> undefined
    # every cell value lies on the [0,1] OE-superiority scale
    pairs = list(A._sample_pairs())[:20]
    for v in A.llm_pairwise_cell(set(pairs)).values():
        assert (0.0 <= v <= 1.0) or np.isnan(v)


if __name__ == '__main__':
    fns = [v for k, v in sorted(globals().items()) if k.startswith('test_') and callable(v)]
    for fn in fns:
        fn(); print(f"ok  {fn.__name__}")
    print(f"\nall {len(fns)} physician-study assertions passed")
